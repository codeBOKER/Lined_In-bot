from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import time

driver = webdriver.Firefox(
    executable_path=r"C:\Program Files (x86)\geckodriver.exe")
driver.get("https://linkedin.com")
wait = WebDriverWait(driver, 20)

# Enter the username and password in the login form
print("please add your email info to log in in the website:- ")
username = input("enter your email: ")
password = input("enter your password: ")

username_field = wait.until(EC.element_to_be_clickable((By.ID, "session_key")))
username_field.send_keys(username)
password_field = wait.until(EC.element_to_be_clickable((By.ID, "session_password")))
password_field.send_keys(password)
# Submit the form to log in
password_field.send_keys(Keys.RETURN)

# get pass the email verification
time.sleep(30)
if driver.find_element(By.ID, "input__email_verification_pin"):
    verfiy_code = input("enter your verfiy code: ")
    verfiy_field = wait.until(EC.element_to_be_clickable((By.ID, "input__email_verification_pin")))
    verfiy_field.send_keys(verfiy_code)

    submit = driver.find_element(By.ID, "email-pin-submit-button").click()
else:
    pass

search = input("Enter what you will search for: ")
search_field = driver.find_element(By.CLASS_NAME, "search-global-typeahead__typeahead").click()
search_input = driver.find_element(By.CLASS_NAME, "search-global-typeahead__input")
search_input.send_keys(search)
search_input.send_keys(Keys.RETURN)

section = driver.find_element(By.CLASS_NAME, "scaffold-layout-toolbar")
filters = section.find_elements(By.CSS_SELECTOR, "button[type='button']")[0].click()

users_link = []
for page in range(1,70):
    container = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.CLASS_NAME, "search-results-container")))
    members = container.find_elements(By.CLASS_NAME, "entity-result")
    print("-"*20, page, "-"*20)
    
    for member in members:
        item = WebDriverWait(member, 30).until(EC.visibility_of_element_located((By.CLASS_NAME, "entity-result__item")))
        div = item.find_element(By.CLASS_NAME, "display-flex")
        a = div.find_element(By.TAG_NAME, "a")
        linkedin_member = "https://www.linkedin.com/search/results/people/headless?origin=SWITCH_SEARCH_VERTICAL&keywords=CEO"
        link = a.get_attribute("href")
        if link != linkedin_member:
            users_link.append(link)
    if page <= 100:
        driver.execute_script("window.scrollBy(0, 2000);")
        Next = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[aria-label='Next']"))).click()   

        
WW = Workbook()
sheets = WW.active
counter = 1
for user_link in users_link:

    driver.get(user_link)
    # finding the name

    profile = WebDriverWait(driver, 40).until(EC.element_to_be_clickable((By.CLASS_NAME, "pv-text-details__left-panel")))
    h1 = profile.find_element(By.TAG_NAME, "h1")
    name = h1.text

    # finding the skills
    skill_div = profile.find_elements(By.TAG_NAME, "div")
    skill = skill_div[1].text


    # finding the followrs
    followers_div = driver.find_element(By.CLASS_NAME, "t-bold")
    followers = followers_div.text

    print(f"{counter}- {name}")
    print(f"skills: {skill}")
    print(f"follower: {followers}")
    print("%"*50)
    
    # set the width of the first column to 20
    sheets.column_dimensions['B'].width = 50
    sheets.column_dimensions['c'].width = 40
    sheets.column_dimensions['d'].width = 200
    
    sheets.cell(column= 1 , row= counter +1, value= name)
    sheets.cell(column= 2 , row= counter +1, value= skill)
    sheets.cell(column= 3 , row= counter +1, value= followers)
    sheets.cell(column= 4 , row= counter +1, value= user_link)
    
    
    WW.save("C:/Users/ab3ad/Desktop/selenium/linkedin_CEO_users.xlsx")
    
    
    counter+=1